from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Any, Iterable

import openpyxl


METADATA_FIRST_CELLS = {"证券代码", "股票代码", "没有单位", "公告日期", "违规事件ID", ""}


def to_float(value: Any, default: float | None = None) -> float | None:
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text or text in {"None", "nan", "NaN", "--"}:
        return default
    text = text.rstrip("%")
    try:
        return float(text)
    except ValueError:
        return default


def stock_code(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).strip()
    if not text or text in {"nan", "None"}:
        return ""
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    match = re.search(r"\d{1,6}", text)
    if not match:
        return ""
    return match.group(0).zfill(6)


def normalize_year(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        try:
            return int(value)
        except (TypeError, ValueError):
            return None
    text = str(value).strip()
    match = re.search(r"\d{4}", text)
    return int(match.group(0)) if match else None


def read_csv_rows(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    encoding = detect_encoding(path)
    with path.open("r", encoding=encoding, newline="") as handle:
        reader = csv.DictReader(handle)
        header = [str(item).strip() for item in (reader.fieldnames or [])]
        rows = []
        for row in reader:
            if is_metadata_row(row, header):
                continue
            rows.append({str(key).strip(): normalize_value(value) for key, value in row.items() if key})
        return rows


def detect_encoding(path: Path) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            with path.open("r", encoding=encoding) as handle:
                handle.read(4096)
            return encoding
        except UnicodeDecodeError:
            continue
    return "utf-8-sig"


def is_metadata_row(row: dict[str, Any], header: Iterable[str] | None = None) -> bool:
    values = list(row.values())
    first = str(values[0] if values else "").strip()
    if first in METADATA_FIRST_CELLS:
        return True
    if "单位" in first and len(first) <= 8:
        return True
    if header and [str(value).strip() for value in values[: len(list(header))]] == list(header):
        return True
    return False


def normalize_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def read_xlsx_rows(path: Path, sheet_name: str | None = None) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook.worksheets[0]
        rows = worksheet.iter_rows(values_only=True)
        header = unique_columns([str(value or "").strip() for value in next(rows, [])])
        result = []
        for values in rows:
            row = dict(zip(header, [normalize_value(value) for value in values]))
            if is_empty_row(row):
                continue
            result.append(row)
        return result
    finally:
        workbook.close()


def unique_columns(columns: list[str]) -> list[str]:
    counts: dict[str, int] = {}
    result = []
    for index, column in enumerate(columns):
        name = column or f"column_{index + 1}"
        counts[name] = counts.get(name, 0) + 1
        if counts[name] > 1:
            name = f"{name}_{counts[name]}"
        result.append(name)
    return result


def is_empty_row(row: dict[str, Any]) -> bool:
    return not any(str(value).strip() for value in row.values() if value is not None)


def latest_item(items: Iterable[dict[str, Any]], year_key: str = "year") -> dict[str, Any] | None:
    valid = [item for item in items if item.get(year_key) is not None]
    if not valid:
        return None
    return sorted(valid, key=lambda item: int(item[year_key]))[-1]


def percent_text(value: Any) -> str:
    number = to_float(value)
    if number is None:
        return "无数据"
    return f"{number:.1f}%"
