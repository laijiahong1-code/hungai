from __future__ import annotations

import csv
import hashlib
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

import openpyxl

from .data import BACKEND_ROOT, DATABASE_ROOT
from .db import JsonDatabase


SUPPORTED_SUFFIXES = {".csv", ".xlsx", ".md"}
METADATA_FIRST_CELLS = {
    "证券代码",
    "股票代码",
    "公告日期",
    "违规事件ID",
    "没有单位",
    "",
}


@dataclass
class RawTable:
    source_id: str
    file_name: str
    sheet_name: str | None
    columns: list[str]
    raw_file: str
    row_count: int


def import_all_sources(
    source_dir: Path | str = BACKEND_ROOT, database_dir: Path | str = DATABASE_ROOT
) -> dict[str, int]:
    source_root = Path(source_dir)
    db_root = Path(database_dir)
    db_root.mkdir(parents=True, exist_ok=True)
    raw_root = db_root / "raw"
    raw_root.mkdir(parents=True, exist_ok=True)

    source_files = [
        path
        for path in sorted(source_root.iterdir(), key=lambda item: item.name)
        if path.is_file() and path.suffix.lower() in SUPPORTED_SUFFIXES
    ]

    catalog = []
    raw_tables: list[RawTable] = []
    policy_documents = []

    for path in source_files:
        source_id = source_id_for(path)
        if path.suffix.lower() == ".md":
            document = read_markdown_document(path, source_id)
            policy_documents.append(document)
            catalog.append(
                {
                    "source_id": source_id,
                    "file_name": path.name,
                    "file_type": "markdown",
                    "size_bytes": path.stat().st_size,
                    "raw_files": [],
                    "row_count": 1,
                    "columns": ["title", "content"],
                }
            )
            continue

        tables = (
            import_csv_raw(path, source_id, raw_root)
            if path.suffix.lower() == ".csv"
            else import_xlsx_raw(path, source_id, raw_root)
        )
        raw_tables.extend(tables)
        catalog.append(
            {
                "source_id": source_id,
                "file_name": path.name,
                "file_type": path.suffix.lower().lstrip("."),
                "size_bytes": path.stat().st_size,
                "raw_files": [table.raw_file for table in tables],
                "row_count": sum(table.row_count for table in tables),
                "columns": tables[0].columns if tables else [],
                "sheets": [
                    {
                        "sheet_name": table.sheet_name,
                        "raw_file": table.raw_file,
                        "row_count": table.row_count,
                        "columns": table.columns,
                    }
                    for table in tables
                ],
            }
        )

    db = JsonDatabase(db_root)
    db.replace_all("source_catalog", catalog)
    db.replace_all("policy_documents", policy_documents)

    derived = build_derived_collections(source_root)
    for collection, records in derived.items():
        db.replace_all(collection, records)

    return {
        "source_count": len(source_files),
        "raw_table_count": len(raw_tables),
        "company_count": len(derived["companies"]),
        "financial_count": len(derived["financials"]),
        "equity_count": len(derived["equity"]),
        "policy_count": len(derived["policies"]),
    }


def source_id_for(path: Path) -> str:
    ascii_stem = re.sub(r"[^a-zA-Z0-9]+", "_", path.stem).strip("_").lower()
    digest = hashlib.sha1(path.name.encode("utf-8")).hexdigest()[:8]
    prefix = ascii_stem[:36] if ascii_stem else "source"
    return f"{prefix}_{digest}"


def sheet_id(name: str) -> str:
    ascii_name = re.sub(r"[^a-zA-Z0-9]+", "_", name).strip("_").lower()
    if ascii_name:
        return ascii_name[:36]
    return hashlib.sha1(name.encode("utf-8")).hexdigest()[:8]


def read_markdown_document(path: Path, source_id: str) -> dict[str, Any]:
    content = path.read_text(encoding="utf-8-sig")
    headings = [line.strip() for line in content.splitlines() if line.lstrip().startswith("#")]
    title = headings[0].lstrip("#").strip() if headings else path.stem
    return {
        "source_id": source_id,
        "source_file": path.name,
        "title": title,
        "headings": headings,
        "content": content,
    }


def import_csv_raw(path: Path, source_id: str, raw_root: Path) -> list[RawTable]:
    encoding = detect_csv_encoding(path)
    raw_file = f"raw/{source_id}.jsonl"
    output_path = raw_root / f"{source_id}.jsonl"
    row_count = 0
    columns: list[str] = []

    with path.open("r", encoding=encoding, newline="") as handle, output_path.open(
        "w", encoding="utf-8"
    ) as output:
        reader = csv.DictReader(handle)
        columns = list(reader.fieldnames or [])
        for row in reader:
            if is_metadata_row(row):
                continue
            record = compact_record(row)
            if not record:
                continue
            output.write(json.dumps(record, ensure_ascii=False) + "\n")
            row_count += 1

    return [RawTable(source_id, path.name, None, columns, raw_file, row_count)]


def import_xlsx_raw(path: Path, source_id: str, raw_root: Path) -> list[RawTable]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    tables = []
    try:
        for worksheet in workbook.worksheets:
            rows = worksheet.iter_rows(values_only=True)
            header = next(rows, None)
            if not header:
                continue
            columns = unique_columns([cell_to_text(value) for value in header])
            if not any(columns):
                continue
            table_id = f"{source_id}__{sheet_id(worksheet.title)}"
            raw_file = f"raw/{table_id}.jsonl"
            output_path = raw_root / f"{table_id}.jsonl"
            row_count = 0
            with output_path.open("w", encoding="utf-8") as output:
                for row_values in rows:
                    row = dict(zip(columns, [cell_to_value(value) for value in row_values]))
                    if is_metadata_row(row):
                        continue
                    record = compact_record(row)
                    if not record:
                        continue
                    output.write(json.dumps(record, ensure_ascii=False) + "\n")
                    row_count += 1
            tables.append(RawTable(source_id, path.name, worksheet.title, columns, raw_file, row_count))
    finally:
        workbook.close()
    return tables


def detect_csv_encoding(path: Path) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            with path.open("r", encoding=encoding) as handle:
                handle.read(4096)
            return encoding
        except UnicodeDecodeError:
            continue
    return "utf-8-sig"


def unique_columns(columns: list[str]) -> list[str]:
    counts: dict[str, int] = {}
    result = []
    for index, column in enumerate(columns):
        name = column.strip() or f"column_{index + 1}"
        counts[name] = counts.get(name, 0) + 1
        if counts[name] > 1:
            name = f"{name}_{counts[name]}"
        result.append(name)
    return result


def is_metadata_row(row: dict[str, Any]) -> bool:
    if not row:
        return True
    first_value = cell_to_text(next(iter(row.values()), ""))
    if first_value in METADATA_FIRST_CELLS:
        return True
    if "单位" in first_value and len(first_value) <= 8:
        return True
    return False


def compact_record(row: dict[str, Any]) -> dict[str, Any]:
    return {
        key: normalize_value(value)
        for key, value in row.items()
        if key and normalize_value(value) not in ("", None)
    }


def normalize_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def cell_to_value(value: Any) -> Any:
    if value is None:
        return ""
    return value


def build_derived_collections(source_root: Path) -> dict[str, list[dict[str, Any]]]:
    companies = load_companies_from_location_workbook(source_root)
    financials = build_financials(source_root, companies)
    equity = build_equity(source_root, companies)
    policies = build_policies(source_root, companies)
    return {
        "companies": companies,
        "financials": financials,
        "equity": equity,
        "policies": policies,
    }


def load_companies_from_location_workbook(source_root: Path) -> list[dict[str, Any]]:
    candidates = list(source_root.glob("*非ST非金融公司所在地*.xlsx"))
    if not candidates:
        return []
    path = candidates[0]
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook["结果明细"] if "结果明细" in workbook.sheetnames else workbook.worksheets[0]
        rows = worksheet.iter_rows(values_only=True)
        header = [cell_to_text(value) for value in next(rows, [])]
        records = []
        for values in rows:
            row = dict(zip(header, [cell_to_value(value) for value in values]))
            code = stock_code(row.get("代码"))
            if not code:
                continue
            records.append(
                {
                    "stock_code": code,
                    "name": cell_to_text(row.get("公司名称")),
                    "short_name": cell_to_text(row.get("简称")),
                    "aliases": [cell_to_text(row.get("简称"))] if row.get("简称") else [],
                    "province": normalize_province_name(cell_to_text(row.get("所在地省份"))),
                    "city": cell_to_text(row.get("所在地城市")),
                    "industry": cell_to_text(row.get("所属行业") or row.get("交易所行业")),
                    "controller": "待补充",
                    "ownership": "待补充",
                    "is_st": False,
                    "is_financial": False,
                }
            )
        return records
    finally:
        workbook.close()


def build_financials(source_root: Path, companies: list[dict[str, Any]]) -> list[dict[str, Any]]:
    codes = {company["stock_code"] for company in companies}
    ratios = latest_csv_by_stock(source_root / "资产负债率、流动比率、速动比率.csv", codes)
    income = latest_csv_by_stock(source_root / "利润表.csv", codes)
    cashflow = latest_csv_by_stock(source_root / "现金流量表（间接法）.csv", codes)
    roe_rows = latest_roe_by_stock(source_root, codes)

    records = []
    for code in sorted(codes):
        ratio = ratios.get(code, {})
        income_row = income.get(code, {})
        cash_row = cashflow.get(code, {})
        roe_row = roe_rows.get(code, {})
        records.append(
            {
                "stock_code": code,
                "revenue": yuan_to_100m(first_number(income_row, ["B001101000", "B001100000"])),
                "net_profit": yuan_to_100m(first_number(cash_row, ["D000101000"])),
                "asset_liability_ratio": ratio_to_percent(first_number(ratio, ["F011201A"])),
                "current_ratio": first_number(ratio, ["F010101A"]),
                "quick_ratio": first_number(ratio, ["F010201A"]),
                "roe": roe_row.get("roe", 0),
                "roe_accper": roe_row.get("roe_accper", ""),
                "cash_flow": yuan_to_100m(first_number(cash_row, ["D000100000"])),
            }
        )
    return records


def build_equity(source_root: Path, companies: list[dict[str, Any]]) -> list[dict[str, Any]]:
    codes = {company["stock_code"] for company in companies}
    pledge = aggregate_pledge(source_root / "股权质押情况.xlsx", codes)
    audit = latest_audit_by_stock(source_root, codes)
    records = []
    for code in sorted(codes):
        pledge_ratio = pledge.get(code, 0.0)
        audit_row = audit.get(code, {})
        records.append(
            {
                "stock_code": code,
                "top_shareholder_ratio": 40.0,
                "pledge_ratio": pledge_ratio,
                "audit_opinion": audit_row.get("audit_opinion", "待补充"),
                "audit_accounting_date": audit_row.get("audit_accounting_date", ""),
                "audit_date": audit_row.get("audit_date", ""),
                "auditor": audit_row.get("auditor", ""),
                "domestic_audit_firm": audit_row.get("domestic_audit_firm", ""),
                "overseas_audit_firm": audit_row.get("overseas_audit_firm", ""),
                "overdue_debt": "待补充",
            }
        )
    return records


def build_policies(source_root: Path, companies: list[dict[str, Any]]) -> list[dict[str, Any]]:
    province_scores = province_policy_scores(source_root)
    records = []
    for company in companies:
        province = company.get("province", "")
        regional_fit = province_scores.get(province, 70.0)
        records.append(
            {
                "stock_code": company["stock_code"],
                "regional_fit": regional_fit,
                "policy_signal": 70.0,
                "positive_reasons": [
                    f"{province or '所在地'}已纳入属地财政与政策数据集",
                    "政策文件和混改样本特征已存入后端数据库",
                ],
                "risk_reasons": ["公司层面的混改事件、控制人和尽调数据仍需继续补充"],
            }
        )
    return records


def latest_csv_by_stock(path: Path, codes: set[str]) -> dict[str, dict[str, Any]]:
    if not path.exists():
        return {}
    latest: dict[str, dict[str, Any]] = {}
    with path.open("r", encoding=detect_csv_encoding(path), newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            if is_metadata_row(row):
                continue
            code = stock_code(row.get("Stkcd"))
            if code not in codes:
                continue
            if row.get("Typrep") and row.get("Typrep") != "A":
                continue
            current = latest.get(code)
            if current is None or cell_to_text(row.get("Accper")) > cell_to_text(current.get("Accper")):
                latest[code] = row
    return latest


def latest_roe_by_stock(source_root: Path, codes: set[str] | None = None) -> dict[str, dict[str, Any]]:
    latest: dict[str, dict[str, Any]] = {}
    for path in roe_csv_candidates(source_root):
        with path.open("r", encoding=detect_csv_encoding(path), newline="") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                if is_metadata_row(row):
                    continue
                code = stock_code(first_text(row, ["Stkcd", "股票代码", "证券代码", "ShortName"]))
                if not code or (codes is not None and code not in codes):
                    continue
                report_type = first_text(row, ["Typrep", "Source"])
                if report_type and report_type != "A" and not looks_like_date(report_type):
                    continue
                roe = to_float(row.get("ROE"))
                if roe is None:
                    continue
                roe_row = {
                    "stock_code": code,
                    "roe": ratio_to_percent(roe),
                    "roe_accper": first_report_date(row),
                }
                current = latest.get(code)
                if current is None or roe_row["roe_accper"] > current.get("roe_accper", ""):
                    latest[code] = roe_row
    return latest


def roe_csv_candidates(source_root: Path) -> list[Path]:
    candidates: list[Path] = []
    seen: set[Path] = set()
    for root in [source_root, source_root.parent]:
        path = root / "ROE.csv"
        if path.exists() and path not in seen:
            candidates.append(path)
            seen.add(path)
    return candidates


def first_report_date(row: dict[str, Any]) -> str:
    for key in ("Accper", "Typrep"):
        value = first_text(row, [key])
        if looks_like_date(value):
            return value
    return ""


def looks_like_date(value: Any) -> bool:
    return bool(re.search(r"\d{4}[-/]\d{1,2}[-/]\d{1,2}", cell_to_text(value)))


def aggregate_pledge(path: Path, codes: set[str]) -> dict[str, float]:
    if not path.exists():
        return {}
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    result: dict[str, float] = {}
    try:
        worksheet = workbook.worksheets[0]
        rows = worksheet.iter_rows(values_only=True)
        header = [cell_to_text(value) for value in next(rows, [])]
        for values in rows:
            row = dict(zip(header, [cell_to_value(value) for value in values]))
            if is_metadata_row(row):
                continue
            code = stock_code(row.get("SCode"))
            if code not in codes:
                continue
            ratio = first_number(row, ["AcPSRaT", "PSRaT", "Ple_ShRat"])
            result[code] = max(result.get(code, 0.0), ratio)
    finally:
        workbook.close()
    return result


def latest_audit_by_stock(source_root: Path, codes: set[str] | None = None) -> dict[str, dict[str, str]]:
    result: dict[str, dict[str, str]] = {}
    for path in audit_workbook_candidates(source_root):
        for code, row in read_audit_workbook(path, codes).items():
            current = result.get(code)
            if current is None or audit_sort_key(row) > audit_sort_key(current):
                result[code] = row
    return result


def audit_workbook_candidates(source_root: Path) -> list[Path]:
    roots = [source_root, source_root.parent]
    names = ["audit_opinions.xlsx", "审计意见.xlsx", "FIN_Audit.xlsx"]
    patterns = ["*audit*opinion*.xlsx", "*审计意见*.xlsx", "*FIN_Audit*.xlsx", "*Audit*.xlsx"]
    candidates: list[Path] = []
    seen: set[Path] = set()
    for root in roots:
        for name in names:
            path = root / name
            if path.exists() and path not in seen:
                candidates.append(path)
                seen.add(path)
        for pattern in patterns:
            for path in sorted(root.glob(pattern), key=lambda item: item.name):
                if path.exists() and path not in seen:
                    candidates.append(path)
                    seen.add(path)
    return candidates


def read_audit_workbook(path: Path, codes: set[str] | None = None) -> dict[str, dict[str, str]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    latest: dict[str, dict[str, str]] = {}
    try:
        worksheet = workbook.worksheets[0]
        rows = worksheet.iter_rows(values_only=True)
        header = unique_columns([cell_to_text(value) for value in next(rows, [])])
        if not header:
            return latest
        for values in rows:
            row = dict(zip(header, [cell_to_value(value) for value in values]))
            if is_metadata_row(row):
                continue
            code = stock_code(first_text(row, ["股票代码", "证券代码", "Stkcd"]))
            if not code or (codes is not None and code not in codes):
                continue
            audit_opinion = first_text(row, ["审计意见", "审计意见类型", "Audittyp", "Adtremark"])
            if not audit_opinion:
                continue
            audit_row = {
                "stock_code": code,
                "audit_opinion": audit_opinion,
                "audit_accounting_date": first_text(row, ["会计截止日期", "Accper"]),
                "audit_date": first_text(row, ["审计日期", "Annodt"]),
                "auditor": first_text(row, ["审计师", "Auditor"]),
                "domestic_audit_firm": first_text(row, ["境内审计事务所", "Dadtunit", "DadtunitID"]),
                "overseas_audit_firm": first_text(row, ["境外审计事务所", "Iadtunit", "IadtunitID"]),
            }
            current = latest.get(code)
            if current is None or audit_sort_key(audit_row) > audit_sort_key(current):
                latest[code] = audit_row
    finally:
        workbook.close()
    return latest


def first_text(row: dict[str, Any], keys: Iterable[str]) -> str:
    for key in keys:
        value = row.get(key)
        text = cell_to_text(value)
        if text:
            return text
    return ""


def audit_sort_key(row: dict[str, str]) -> tuple[str, str]:
    return (row.get("audit_accounting_date", ""), row.get("audit_date", ""))


def province_policy_scores(source_root: Path) -> dict[str, float]:
    path = source_root / "china_statistical_2024_province_fiscal_gdp_indicators.xlsx"
    if not path.exists():
        return {}
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    scores: dict[str, float] = {}
    try:
        worksheet = workbook["Data_2024"] if "Data_2024" in workbook.sheetnames else workbook.worksheets[0]
        rows = worksheet.iter_rows(values_only=True)
        header = [cell_to_text(value) for value in next(rows, [])]
        for values in rows:
            row = dict(zip(header, [cell_to_value(value) for value in values]))
            province = normalize_province_name(cell_to_text(row.get("city_name")))
            pressure = first_number(row, ["fiscal_pressure_pct"])
            if province:
                scores[province] = round(max(45.0, min(95.0, 90.0 - pressure * 0.5)), 1)
    finally:
        workbook.close()
    return scores


def stock_code(value: Any) -> str:
    text = cell_to_text(value)
    if not text:
        return ""
    text = text.split(".")[0]
    if text.isdigit():
        return text.zfill(6)
    match = re.search(r"\d{6}", text)
    return match.group(0) if match else ""


def first_number(row: dict[str, Any], keys: Iterable[str]) -> float:
    for key in keys:
        value = row.get(key)
        number = to_float(value)
        if number is not None:
            return number
    return 0.0


def to_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def yuan_to_100m(value: float) -> float:
    return round(value / 100000000, 2)


def ratio_to_percent(value: float) -> float:
    if value <= 2:
        return round(value * 100, 2)
    return round(value, 2)


def normalize_province_name(value: str) -> str:
    if not value:
        return ""
    direct = {"北京": "北京市", "上海": "上海市", "天津": "天津市", "重庆": "重庆市"}
    if value in direct:
        return direct[value]
    if value.endswith(("省", "市", "自治区")):
        return value
    return f"{value}省"


if __name__ == "__main__":
    result = import_all_sources()
    print(json.dumps(result, ensure_ascii=False, indent=2))
