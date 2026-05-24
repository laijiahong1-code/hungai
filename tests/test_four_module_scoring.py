import csv
import json
from pathlib import Path

import openpyxl

from backend.app.scoring.engine import ScoringEngine


def write_csv(path: Path, header: list[str], rows: list[list[object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(header)
        writer.writerow(header)
        writer.writerow(["没有单位"] * len(header))
        writer.writerows(rows)


def write_workbook(path: Path, sheet_name: str, header: list[str], rows: list[list[object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(header)
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def write_source_root(root: Path) -> None:
    finance_dir = root / "财务引资潜力数据" / "智能财务机器人数据"
    write_csv(
        finance_dir / "资产负债表.csv",
        [
            "Stkcd",
            "Accper",
            "A001000000",
            "A001100000",
            "A002100000",
            "A003105000",
            "A002000000",
            "A002101000",
            "A002201000",
            "A002203000",
            "A0f2104000",
            "A0b2102000",
            "A0b2103000",
        ],
        [["600001", "2024-12-31", 100, 60, 30, 20, 50, 10, 0, 0, 0, 0, 0]],
    )
    write_csv(
        finance_dir / "利润表.csv",
        ["Stkcd", "Accper", "B001300000", "B001100000", "B002000101", "B002000000"],
        [
            ["600001", "2022-12-31", 8, 100, 100, 100],
            ["600001", "2023-12-31", 9, 110, 121, 121],
            ["600001", "2024-12-31", 10, 120, 146.41, 146.41],
        ],
    )
    write_csv(
        finance_dir / "现金流量表（直接法）.csv",
        ["Stkcd", "Accper", "C001000000"],
        [["600001", "2024-12-31", 20]],
    )
    write_csv(
        finance_dir / "资产负债率、流动比率、速动比率.csv",
        ["Stkcd", "Accper", "F011201A", "F010101A", "F010201A"],
        [["600001", "2024-12-31", 0.5, 1.2, 1.0]],
    )
    write_csv(
        finance_dir / "红利分配文件.csv",
        ["Stkcd", "Accper", "Btperdiv", "Atperdiv", "Numdiv", "DistributionBaseShares", "Price1"],
        [
            ["600001", "2022-12-31", 0.1, 0, 0, 100, 1],
            ["600001", "2023-12-31", 0.1, 0, 0, 100, 1],
            ["600001", "2024-12-31", 0.1, 0, 0, 100, 1],
        ],
    )

    governance_dir = root / "企业股权评分" / "企业股权评分"
    data_dir = governance_dir / "data"
    shareholder_rows = [
        ["中文行", "", "", ""],
        ["单位行", "", "", ""],
        ["说明行", "", "", ""],
    ]
    shares = [30, 15, 4, 4, 4, 3, 3, 3, 2, 2]
    shareholder_rows.extend(
        ["600001", "2024-12-31", index, ratio] for index, ratio in enumerate(shares, start=1)
    )
    write_workbook(data_dir / "十大股东.xlsx", "Sheet1", ["Stkcd", "Reptdt", "S0201b", "S0301b"], shareholder_rows)
    write_workbook(data_dir / "股权质押.xlsx", "Sheet1", ["股票代码", "质押比例"], [["600001", 0.05]])
    write_workbook(data_dir / "审计意见.xlsx", "Sheet1", ["股票代码", "审计意见"], [["600001", "标准无保留意见"]])
    write_workbook(data_dir / "违规信息.xlsx", "Sheet1", ["股票代码", "违规类型"], [["600002", "一般违规"]])
    write_workbook(data_dir / "龙头企业.xlsx", "Sheet1", ["股票代码"], [["600001"]])
    write_workbook(
        governance_dir / "result" / "企业股权最终评分.xlsx",
        "Sheet1",
        ["股票代码", "截止日期", "最终总分", "年份"],
        [
            ["600001", "2022-12-31", 16.0, 2022],
            ["600001", "2023-12-31", 18.0, 2023],
            ["600001", "2024-03-31", 19.0, 2024],
            ["600001", "2024-12-31", 20.0, 2024],
            ["600001", "2025-09-30", 21.25, 2025],
            ["600002", "2025-12-31", 10.0, 2025],
        ],
    )

    region_dir = root / "区域国资匹配度" / "区域国资匹配度——代码初稿"
    company_rows = [["600001", "甲能源", "甲能源集团", "能源", "江西省", "南昌市"]]
    for index in range(2, 11):
        company_rows.append([str(600000 + index), f"样本{index}", f"样本{index}公司", "能源", "江西省", "南昌市"])
    write_workbook(
        region_dir / "A股主板非ST非金融公司所在地_2026-05-07.xlsx",
        "结果明细",
        ["代码", "简称", "公司名称", "所属行业", "所在地省份", "所在地城市"],
        company_rows,
    )
    workbook = openpyxl.Workbook()
    detail = workbook.active
    detail.title = "Province_Detail_2023"
    detail.append(["province_name", "city_name", "revenue_expenditure_ratio_pct"])
    detail.append(["江西省", "南昌市", 100])
    summary = workbook.create_sheet("Province_Summary_2023")
    summary.append(["province_name", "revenue_expenditure_ratio_pct"])
    summary.append(["江西省", 100])
    fiscal_city = region_dir / "china_prefecture_city_fiscal_gdp_indicators_2023_by_province(2).xlsx"
    fiscal_city.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(fiscal_city)
    write_workbook(
        region_dir / "china_statistical_2024_province_fiscal_gdp_indicators(1).xlsx",
        "Data_2024",
        ["city_name", "revenue_expenditure_ratio_pct", "gdp_current_100m_yuan"],
        [["江西省", 100, 1000]],
    )
    (region_dir / "地方负债率（省级）.xls").write_text(
        "<table><tr><td>地区</td><td>2024年</td></tr><tr><td>江西省</td><td>600</td></tr></table>",
        encoding="utf-8",
    )
    write_workbook(
        region_dir / "企业混改程度表.xlsx",
        "GA_StateOwnedMixedDegree",
        ["Symbol", "是否发生混改", "EndDate"],
        [[str(600000 + index), 1, "2025-12-31"] for index in range(1, 11)],
    )
    (region_dir / "industry_match_cache.json").write_text(
        json.dumps({"能源|江西省|南昌市": "省级重点产业"}, ensure_ascii=False),
        encoding="utf-8",
    )

    mixed_dir = root / "混改程度评分" / "国有上市公司混改程度表105735375(仅供厦门大学使用)"
    write_csv(
        mixed_dir / "GA_StateOwnedMixedDegree.csv",
        [
            "InstitutionID",
            "Symbol",
            "ShortName",
            "EndDate",
            "MixedEquityStructureOrNOT",
            "NSttOwnedShrhlderRatioSum",
            "NSttOwnedShrhlderPartic",
            "EquityStructureDiversity",
            "EquityStrucEntropyIndex",
            "EquityBalance",
            "OwnershipConcentration",
            "EquityStrucHerfindalIndex",
            "EquityIntegration",
        ],
        [[1, "600001", "甲能源", "2025-12-31", 1, 35, 0.7, 6, 1.0, 1.0, 55, 0.35, 1.0]],
    )


def test_scoring_engine_recomputes_four_modules_from_raw_sources(tmp_path):
    write_source_root(tmp_path)

    result = ScoringEngine(tmp_path).score_company("600001")

    assert result["modules"] == {
        "finance": 98.0,
        "equity": 100.0,
        "region": 85.0,
        "mixed": 78.8,
    }
    assert result["raw_scores"] == {
        "finance": {"score": 49.0, "max": 50.0},
        "equity": {"score": 25.0, "max": 25.0},
        "region": {"score": 17.0, "max": 20.0},
        "mixed": {"score": 78.8, "max": 100.0},
    }
    assert result["totalScore"] == 93.0
    assert result["potentialLevel"] == "高潜力"
    assert result["vetoReasons"] == []
    assert result["module_details"]["equity"]["evidence"][0]["max"] == 5.0


def test_scoring_engine_exposes_governance_trend_from_final_scores(tmp_path):
    write_source_root(tmp_path)

    result = ScoringEngine(tmp_path).score_company("600001")

    assert result["governanceTrend"] == [
        {"year": 2023, "score": 72.0, "rawScore": 18.0, "date": "2023-12-31"},
        {"year": 2024, "score": 80.0, "rawScore": 20.0, "date": "2024-12-31"},
        {"year": 2025, "score": 85.0, "rawScore": 21.25, "date": "2025-09-30"},
    ]


def test_scoring_engine_returns_zero_scores_when_raw_company_is_missing(tmp_path):
    write_source_root(tmp_path)

    result = ScoringEngine(tmp_path).score_company("600999")

    assert result["modules"] == {"finance": 0.0, "equity": 0.0, "region": 0.0, "mixed": 0.0}
    assert result["governanceTrend"] == []
    assert result["totalScore"] == 0.0
    assert result["potentialLevel"] == "低潜力"
